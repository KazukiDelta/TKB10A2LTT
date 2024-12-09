from openpyxl import load_workbook
#file được làm bởi Kazuki Delta khi đang chơi đá
#Credit: Delta

file_goc = 'TKBCHINH.xlsx'
wb_goc = load_workbook(file_goc)
sheet_goc = wb_goc['DATASHEET']

F4 = sheet_goc['F4'].value
F5 = sheet_goc['F5'].value
F6 = sheet_goc['F6'].value
F7 = sheet_goc['F7'].value
F8 = sheet_goc['F8'].value

F10 = sheet_goc['F10'].value
F11 = sheet_goc['F11'].value
F12 = sheet_goc['F12'].value
F13 = sheet_goc['F13'].value

F14 = sheet_goc['F14'].value
F15 = sheet_goc['F15'].value
F16 = sheet_goc['F16'].value
F17 = sheet_goc['F17'].value
F18 = sheet_goc['F18'].value

F20 = sheet_goc['F20'].value
F21 = sheet_goc['F21'].value
F22 = sheet_goc['F22'].value
F23 = sheet_goc['F23'].value

F24 = sheet_goc['F24'].value
F25 = sheet_goc['F25'].value
F26 = sheet_goc['F26'].value
F27 = sheet_goc['F27'].value
F28 = sheet_goc['F28'].value

F30 = sheet_goc['F30'].value
F31 = sheet_goc['F31'].value
F32 = sheet_goc['F32'].value
F33 = sheet_goc['F33'].value

F34 = sheet_goc['F34'].value
F35 = sheet_goc['F35'].value
F36 = sheet_goc['F36'].value
F37 = sheet_goc['F37'].value
F38 = sheet_goc['F38'].value

F40 = sheet_goc['F40'].value
F41 = sheet_goc['F41'].value
F42 = sheet_goc['F42'].value
F43 = sheet_goc['F43'].value

F44 = sheet_goc['F44'].value
F45 = sheet_goc['F45'].value
F46 = sheet_goc['F46'].value
F47 = sheet_goc['F47'].value
F48 = sheet_goc['F48'].value

F50 = sheet_goc['F50'].value
F51 = sheet_goc['F51'].value
F52 = sheet_goc['F52'].value
F53 = sheet_goc['F53'].value

F54 = sheet_goc['F54'].value
F55 = sheet_goc['F55'].value
F56 = sheet_goc['F56'].value
F57 = sheet_goc['F57'].value
F58 = sheet_goc['F58'].value

F60 = sheet_goc['F60'].value
F61 = sheet_goc['F61'].value
F62 = sheet_goc['F62'].value
F63 = sheet_goc['F63'].value

file_moi = 'TKB.xlsx'
wb_moi = load_workbook(file_moi)
ten_sheet = 'Sheet1'
if ten_sheet in wb_moi.sheetnames:
    sheet = wb_moi[ten_sheet]
else:
    sheet = wb_moi.create_sheet(ten_sheet)

sheet['C3'] = F4
sheet['C4'] = F5
sheet['C6'] = F6
sheet['C7'] = F7
sheet['C8'] = F8

sheet['C10'] = F10
sheet['C11'] = F11
sheet['C13'] = F12
sheet['C14'] = F13

sheet['D3'] = F14
sheet['D4'] = F15
sheet['D6'] = F16
sheet['D7'] = F17
sheet['D8'] = F18

sheet['D10'] = F20
sheet['D11'] = F21
sheet['D13'] = F22
sheet['D14'] = F23

sheet['E3'] = F24
sheet['E4'] = F25
sheet['E6'] = F26
sheet['E7'] = F27
sheet['E8'] = F28

sheet['E10'] = F30
sheet['E11'] = F31
sheet['E13'] = F32
sheet['E14'] = F33

sheet['F3'] = F34
sheet['F4'] = F35
sheet['F6'] = F36
sheet['F7'] = F37
sheet['F8'] = F38

sheet['F10'] = F40
sheet['F11'] = F41
sheet['F13'] = F42
sheet['F14'] = F43

sheet['G3'] = F44
sheet['G4'] = F45
sheet['G6'] = F46
sheet['G7'] = F47
sheet['G8'] = F48

sheet['G10'] = F50
sheet['G11'] = F51
sheet['G13'] = F52
sheet['G14'] = F53

sheet['H3'] = F54
sheet['H4'] = F55
sheet['H6'] = F56
sheet['H7'] = F57
sheet['H8'] = F58

sheet['H10'] = F60
sheet['H11'] = F61
sheet['H13'] = F62
sheet['H14'] = F63

wb_moi.save(file_moi)
print("It's done you lazy bitch")