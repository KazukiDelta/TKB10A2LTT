from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

#file được làm bởi Kazuki Delta khi đang chơi đá
#Credit: Delta

file_goc = 'TKBCHINH.xlsx'
wb_goc = load_workbook(file_goc)
sheet_goc = wb_goc['DATASHEET']

F4 = sheet_goc['F4'].value
F5 = sheet_goc['F5'].value
F6 = sheet_goc['F6'].value
F7 = sheet_goc['F7'].value
F8 = sheet_goc['F8'].value

F10 = sheet_goc['F10'].value
F11 = sheet_goc['F11'].value
F12 = sheet_goc['F12'].value
F13 = sheet_goc['F13'].value

F14 = sheet_goc['F14'].value
F15 = sheet_goc['F15'].value
F16 = sheet_goc['F16'].value
F17 = sheet_goc['F17'].value
F18 = sheet_goc['F18'].value

F20 = sheet_goc['F20'].value
F21 = sheet_goc['F21'].value
F22 = sheet_goc['F22'].value
F23 = sheet_goc['F23'].value

F24 = sheet_goc['F24'].value
F25 = sheet_goc['F25'].value
F26 = sheet_goc['F26'].value
F27 = sheet_goc['F27'].value
F28 = sheet_goc['F28'].value

F30 = sheet_goc['F30'].value
F31 = sheet_goc['F31'].value
F32 = sheet_goc['F32'].value
F33 = sheet_goc['F33'].value

F34 = sheet_goc['F34'].value
F35 = sheet_goc['F35'].value
F36 = sheet_goc['F36'].value
F37 = sheet_goc['F37'].value
F38 = sheet_goc['F38'].value

F40 = sheet_goc['F40'].value
F41 = sheet_goc['F41'].value
F42 = sheet_goc['F42'].value
F43 = sheet_goc['F43'].value

F44 = sheet_goc['F44'].value
F45 = sheet_goc['F45'].value
F46 = sheet_goc['F46'].value
F47 = sheet_goc['F47'].value
F48 = sheet_goc['F48'].value

F50 = sheet_goc['F50'].value
F51 = sheet_goc['F51'].value
F52 = sheet_goc['F52'].value
F53 = sheet_goc['F53'].value

F54 = sheet_goc['F54'].value
F55 = sheet_goc['F55'].value
F56 = sheet_goc['F56'].value
F57 = sheet_goc['F57'].value
F58 = sheet_goc['F58'].value

F60 = sheet_goc['F60'].value
F61 = sheet_goc['F61'].value
F62 = sheet_goc['F62'].value
F63 = sheet_goc['F63'].value

file_moi = 'TKB.xlsx'
wb_moi = load_workbook(file_moi)
ten_sheet = 'Sheet1'
if ten_sheet in wb_moi.sheetnames:
    sheet = wb_moi[ten_sheet]
else:
    sheet = wb_moi.create_sheet(ten_sheet)

sheet['C3'] = F4
sheet['C4'] = F5
sheet['C6'] = F6
sheet['C7'] = F7
sheet['C8'] = F8

sheet['C10'] = F10
sheet['C11'] = F11
sheet['C13'] = F12
sheet['C14'] = F13

sheet['D3'] = F14
sheet['D4'] = F15
sheet['D6'] = F16
sheet['D7'] = F17
sheet['D8'] = F18

sheet['D10'] = F20
sheet['D11'] = F21
sheet['D13'] = F22
sheet['D14'] = F23

sheet['E3'] = F24
sheet['E4'] = F25
sheet['E6'] = F26
sheet['E7'] = F27
sheet['E8'] = F28

sheet['E10'] = F30
sheet['E11'] = F31
sheet['E13'] = F32
sheet['E14'] = F33

sheet['F3'] = F34
sheet['F4'] = F35
sheet['F6'] = F36
sheet['F7'] = F37
sheet['F8'] = F38

sheet['F10'] = F40
sheet['F11'] = F41
sheet['F13'] = F42
sheet['F14'] = F43

sheet['G3'] = F44
sheet['G4'] = F45
sheet['G6'] = F46
sheet['G7'] = F47
sheet['G8'] = F48

sheet['G10'] = F50
sheet['G11'] = F51
sheet['G13'] = F52
sheet['G14'] = F53

sheet['H3'] = F54
sheet['H4'] = F55
sheet['H6'] = F56
sheet['H7'] = F57
sheet['H8'] = F58

sheet['H10'] = F60
sheet['H11'] = F61
sheet['H13'] = F62
sheet['H14'] = F63

sheet['C3'].fill = sheet_goc['F4'].fill.copy()
sheet['C4'].fill = sheet_goc['F5'].fill.copy()
sheet['C6'].fill = sheet_goc['F6'].fill.copy()
sheet['C7'].fill = sheet_goc['F7'].fill.copy()
sheet['C8'].fill = sheet_goc['F8'].fill.copy()

sheet['C10'].fill = sheet_goc['F10'].fill.copy()
sheet['C11'].fill = sheet_goc['F11'].fill.copy()
sheet['C13'].fill = sheet_goc['F12'].fill.copy()
sheet['C14'].fill = sheet_goc['F13'].fill.copy()

sheet['D3'].fill = sheet_goc['F14'].fill.copy()
sheet['D4'].fill = sheet_goc['F15'].fill.copy()
sheet['D6'].fill = sheet_goc['F16'].fill.copy()
sheet['D7'].fill = sheet_goc['F17'].fill.copy()
sheet['D8'].fill = sheet_goc['F18'].fill.copy()

sheet['D10'].fill = sheet_goc['F20'].fill.copy()
sheet['D11'].fill = sheet_goc['F21'].fill.copy()
sheet['D13'].fill = sheet_goc['F22'].fill.copy()
sheet['D14'].fill = sheet_goc['F23'].fill.copy()

sheet['E3'].fill = sheet_goc['F24'].fill.copy()
sheet['E4'].fill = sheet_goc['F25'].fill.copy()
sheet['E6'].fill = sheet_goc['F26'].fill.copy()
sheet['E7'].fill = sheet_goc['F27'].fill.copy()
sheet['E8'].fill = sheet_goc['F28'].fill.copy()

sheet['E10'].fill = sheet_goc['F30'].fill.copy()
sheet['E11'].fill = sheet_goc['F31'].fill.copy()
sheet['E13'].fill = sheet_goc['F32'].fill.copy()
sheet['E14'].fill = sheet_goc['F33'].fill.copy()

sheet['F3'].fill = sheet_goc['F34'].fill.copy()
sheet['F4'].fill = sheet_goc['F35'].fill.copy()
sheet['F6'].fill = sheet_goc['F36'].fill.copy()
sheet['F7'].fill = sheet_goc['F37'].fill.copy()
sheet['F8'].fill = sheet_goc['F38'].fill.copy()

sheet['F10'].fill = sheet_goc['F40'].fill.copy()
sheet['F11'].fill = sheet_goc['F41'].fill.copy()
sheet['F13'].fill = sheet_goc['F42'].fill.copy()
sheet['F14'].fill = sheet_goc['F43'].fill.copy()

sheet['G3'].fill = sheet_goc['F44'].fill.copy()
sheet['G4'].fill = sheet_goc['F45'].fill.copy()
sheet['G6'].fill = sheet_goc['F46'].fill.copy()
sheet['G7'].fill = sheet_goc['F47'].fill.copy()
sheet['G8'].fill = sheet_goc['F48'].fill.copy()

sheet['G10'].fill = sheet_goc['F50'].fill.copy()
sheet['G11'].fill = sheet_goc['F51'].fill.copy()
sheet['G13'].fill = sheet_goc['F52'].fill.copy()
sheet['G14'].fill = sheet_goc['F53'].fill.copy()

sheet['H3'].fill = sheet_goc['F54'].fill.copy()
sheet['H4'].fill = sheet_goc['F55'].fill.copy()
sheet['H6'].fill = sheet_goc['F56'].fill.copy()
sheet['H7'].fill = sheet_goc['F57'].fill.copy()
sheet['H8'].fill = sheet_goc['F58'].fill.copy()

sheet['H10'].fill = sheet_goc['F60'].fill.copy()
sheet['H11'].fill = sheet_goc['F61'].fill.copy()
sheet['H13'].fill = sheet_goc['F62'].fill.copy()
sheet['H14'].fill = sheet_goc['F63'].fill.copy()

sheet.merge_cells('C2:H2')
sheet['C2'] = "S"

sheet.merge_cells('C5:H5')
sheet['C5'] = "RA CHƠI"

sheet.merge_cells('C9:H9')
sheet['C9'] = "C"

sheet.merge_cells('C12:H12')
sheet['C12'] = "GIẢI LAO 10P"

thoi_gian_sang = [
    "Thời Gian",
    "6h50-7h00",
    "7h00-7h45",
    "7h50-8h35",
    "8h35-9h05",
    "9h05-9h50",
    "9h55-10h40",
    "10h45-11h30"
]
for row, value in enumerate(thoi_gian_sang, start=1):
    sheet[f'B{row}'] = value

thoi_gian_chieu = [
    "13h40-14h25",
    "14h30-15h15",
    "15h15-15h25",
    "15h25-16h10",
    "16h15-17h00"
]

for row, value in enumerate(thoi_gian_chieu, start=10):
    sheet[f'B{row}'] = value

tiet = [
    "Thứ",
    "Buổi",
    "Tiết 1",
    "Tiết 2",
    "",
    "Tiết 3",
    "Tiết 4",
    "Tiết 5",
    "Buổi",
    "Tiết 2",
    "Tiết 3",
    "",
    "Tiết 4",
    "Tiết 5"
]

for row, value in enumerate(tiet, start=1):
    sheet[f'A{row}'] = value
sheet['C1'] = 'Thứ 2'
sheet['D1'] = 'Thứ 3'
sheet['E1'] = 'Thứ 4'
sheet['F1'] = 'Thứ 5'
sheet['G1'] = 'Thứ 6'
sheet['H1'] = 'Thứ 7'

mau_xanh = PatternFill(start_color="42e3f5", end_color="42e3f5", fill_type="solid")  # Xanh lục lam
mau_vang = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")  # Vàng

red = PatternFill(start_color="f52c2c", end_color="f52c2c", fill_type='solid')  # Đỏ
orange = PatternFill(start_color="ff8400", end_color="ff8400", fill_type='solid')  # Cam
yellow = PatternFill(start_color="fae505", end_color="fae505", fill_type='solid')  # Vàng
green = PatternFill(start_color="6fff00", end_color="6fff00", fill_type='solid')  # Xanh lá
blue = PatternFill(start_color="0559f5", end_color="0559f5", fill_type='solid')  # Xanh dương
purple = PatternFill(start_color="7905f5", end_color="7905f5", fill_type='solid')  # Tím


sheet['C1'].fill = red
sheet['D1'].fill = orange
sheet['E1'].fill = yellow
sheet['F1'].fill = green
sheet['G1'].fill = blue
sheet['H1'].fill = purple


o_mau_xanh = ['A1', 'A3', 'A4', 'A6', 'A7', 'A8', 'A10', 'A11', 'A13', 'A14']
o_mau_vang = ['A2', 'A5', 'A9', 'A12']

for cell in o_mau_xanh:
    sheet[cell].fill = mau_xanh

for cell in o_mau_vang:
    sheet[cell].fill = mau_vang

thick_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
times_new_roman = Font(name='Times New Roman', size=9)
alignment_center = Alignment(horizontal='center', vertical='center')
font_bold = Font(bold=True)

for row in sheet.iter_rows(min_col=1, max_col=8):
    for cell in row:
        cell.border = thick_border
        cell.font = times_new_roman
        cell.alignment = alignment_center
        cell.font = font_bold 
  
for col in range(ord('C'), ord('H') + 1):
    sheet.column_dimensions[chr(col)].width = 30
for row in range(1,15):
    sheet.row_dimensions[row].height = 25
sheet.column_dimensions['B'].width = 15
wb_moi.save(file_moi)
print("It's done you lazy bitch")